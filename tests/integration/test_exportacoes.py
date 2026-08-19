"""
O número de sócio precisa sair nas três saídas de dados que restaram: CSV,
Excel e busca do admin. (Havia uma quarta, o payload do webhook, removida
junto com a integração na migration 008.)
"""

import csv
import io


async def _criar_associado(
    db_session, nome, cpf, numero, telefone, modalidades=None, outros=None
):
    from app.models import (
        Associado,
        AssociadoDepartamento,
        Candidato,
        Departamento,
        Preferencia,
        Resposta,
    )

    candidato = Candidato(nome="Fulano", apelido="Fu", ativo=True)
    db_session.add(candidato)
    await db_session.commit()
    await db_session.refresh(candidato)

    associado = Associado(
        nome=nome,
        cpf=cpf,
        numero_socio=numero,
        telefone=telefone,
        departamento_outros=outros,
        aceite_lgpd=True,
    )
    db_session.add(associado)
    await db_session.commit()
    await db_session.refresh(associado)

    db_session.add(Resposta(associado_id=associado.id, candidato_id=candidato.id))
    db_session.add(
        Preferencia(associado_id=associado.id, candidato_preferido_id=candidato.id)
    )

    for ordem, nome_mod in enumerate(modalidades or [], start=1):
        dep = Departamento(nome=nome_mod, ordem=ordem, ativo=True)
        db_session.add(dep)
        await db_session.commit()
        await db_session.refresh(dep)
        db_session.add(
            AssociadoDepartamento(associado_id=associado.id, departamento_id=dep.id)
        )

    await db_session.commit()
    return associado


async def _preparar_voto_export(
    client, db_session, valid_cpf, numero_socio, read_otp_code, telefone
):
    """Igual ao de test_departamentos.py — repetido aqui de propósito para
    que cada arquivo de teste seja legível sozinho."""
    from app.models import Candidato

    candidato = Candidato(nome="Fulano", apelido="Fu", ativo=True)
    db_session.add(candidato)
    await db_session.commit()
    await db_session.refresh(candidato)

    res = await client.post(
        "/api/survey/register",
        json={
            "nome": "Socio Export",
            "cpf": valid_cpf(),
            "telefone": telefone,
            "numero_socio": numero_socio(),
            "titular": True,
            "recaptcha_token": "",
        },
    )
    token = res.json()["session_token"]
    codigo = read_otp_code(telefone)
    await client.post(
        "/api/survey/verify-otp",
        json={"session_token": token, "telefone": telefone, "codigo": codigo},
    )
    return token, candidato.id


class TestExportacaoCSV:
    async def test_csv_tem_coluna_de_numero_de_socio(
        self, client, db_session, admin_token, valid_cpf
    ):
        await _criar_associado(db_session, "Socio CSV", valid_cpf(), "0042", "11944440001")

        res = await client.get(
            "/api/admin/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.status_code == 200

        linhas = list(csv.reader(io.StringIO(res.text)))
        assert "Nº Sócio" in linhas[0]

        indice = linhas[0].index("Nº Sócio")
        assert linhas[1][indice] == "0042"


class TestExportacaoExcel:
    async def test_excel_tem_coluna_de_numero_de_socio(
        self, client, db_session, admin_token, valid_cpf
    ):
        from openpyxl import load_workbook

        await _criar_associado(db_session, "Socio XLS", valid_cpf(), "0043", "11944440002")

        res = await client.get(
            "/api/admin/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.status_code == 200

        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        cabecalho = [c.value for c in ws[1]]
        assert "Nº Sócio" in cabecalho

        indice = cabecalho.index("Nº Sócio")
        assert ws[2][indice].value == "0043"


class TestBuscaDoAdmin:
    async def test_busca_devolve_numero_de_socio(
        self, client, db_session, admin_token, valid_cpf
    ):
        cpf = valid_cpf()
        await _criar_associado(db_session, "Socio Busca", cpf, "0044", "11944440003")

        res = await client.get(
            f"/api/admin/search?cpf={cpf}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.status_code == 200
        assert res.json()[0]["numero_socio"] == "0044"


class TestModalidadesNasExportacoes:
    async def test_csv_tem_as_duas_colunas(
        self, client, db_session, admin_token, valid_cpf
    ):
        await _criar_associado(
            db_session,
            "Socio Mod",
            valid_cpf(),
            "0050",
            "11922220001",
            modalidades=["Natação", "Sauna"],
            outros=None,
        )

        res = await client.get(
            "/api/admin/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        linhas = list(csv.reader(io.StringIO(res.text)))
        assert "Modalidades" in linhas[0]
        assert "Outros (descrição)" in linhas[0]
        assert linhas[1][linhas[0].index("Modalidades")] == "Natação, Sauna"

    async def test_excel_tem_as_duas_colunas(
        self, client, db_session, admin_token, valid_cpf
    ):
        from openpyxl import load_workbook

        await _criar_associado(
            db_session,
            "Socio Mod XLS",
            valid_cpf(),
            "0051",
            "11922220002",
            modalidades=["Judô"],
            outros="Xadrez",
        )

        res = await client.get(
            "/api/admin/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        ws = load_workbook(io.BytesIO(res.content)).active
        cabecalho = [c.value for c in ws[1]]
        assert ws[2][cabecalho.index("Modalidades")].value == "Judô"
        assert ws[2][cabecalho.index("Outros (descrição)")].value == "Xadrez"

    async def test_busca_do_admin_devolve_as_modalidades(
        self, client, db_session, admin_token, valid_cpf
    ):
        cpf = valid_cpf()
        await _criar_associado(
            db_session,
            "Socio Busca Mod",
            cpf,
            "0052",
            "11922220003",
            modalidades=["Piscina"],
            outros=None,
        )

        res = await client.get(
            f"/api/admin/search?cpf={cpf}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.json()[0]["departamentos"] == ["Piscina"]
        assert res.json()[0]["departamento_outros"] == ""

    async def test_csv_respeita_a_ordem_e_nao_o_nome(
        self, client, db_session, admin_token, valid_cpf
    ):
        """Mesmo motivo do teste equivalente em test_departamentos.py: sob a
        collation do banco, ordenar por nome devolveria outra sequência.
        "Zumba" (ordem 1) antes de "Atletismo" (ordem 2) torna a divergência
        óbvia — uma regressão para sort por nome devolveria a ordem trocada."""
        await _criar_associado(
            db_session,
            "Socio Mod Ordem CSV",
            valid_cpf(),
            "0053",
            "11922220006",
            modalidades=["Zumba", "Atletismo"],
            outros=None,
        )

        res = await client.get(
            "/api/admin/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        linhas = list(csv.reader(io.StringIO(res.text)))
        assert linhas[1][linhas[0].index("Modalidades")] == "Zumba, Atletismo"

    async def test_excel_respeita_a_ordem_e_nao_o_nome(
        self, client, db_session, admin_token, valid_cpf
    ):
        """Mesmo motivo do teste equivalente em test_departamentos.py."""
        from openpyxl import load_workbook

        await _criar_associado(
            db_session,
            "Socio Mod Ordem XLS",
            valid_cpf(),
            "0054",
            "11922220007",
            modalidades=["Zumba", "Atletismo"],
            outros=None,
        )

        res = await client.get(
            "/api/admin/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        ws = load_workbook(io.BytesIO(res.content)).active
        cabecalho = [c.value for c in ws[1]]
        assert ws[2][cabecalho.index("Modalidades")].value == "Zumba, Atletismo"

    async def test_busca_do_admin_respeita_a_ordem_e_nao_o_nome(
        self, client, db_session, admin_token, valid_cpf
    ):
        """Mesmo motivo do teste equivalente em test_departamentos.py."""
        cpf = valid_cpf()
        await _criar_associado(
            db_session,
            "Socio Mod Ordem Busca",
            cpf,
            "0055",
            "11922220008",
            modalidades=["Zumba", "Atletismo"],
            outros=None,
        )

        res = await client.get(
            f"/api/admin/search?cpf={cpf}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert res.json()[0]["departamentos"] == ["Zumba", "Atletismo"]
